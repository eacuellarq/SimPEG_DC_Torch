"""
Benchmark: Solvers con modelos de resistividad DC 2D
=====================================================
Compara SuperLU (CPU), Pardiso (CPU), PCG-GPU (CUDA), Dense-GPU (torch.linalg.solve)
con modelos de resistividad de distintos tamanos.

Genera 3 graficas: forward time, backward time, memoria.

Ejecutar desde DC_torch/:
    cd DC_torch
    python ../examples/benchmark_solvers_dc.py
"""

import os
os.environ["KMP_DUPLICATE_LIB_OK"] = "TRUE"

import gc
import time
import numpy as np
import torch
import psutil
import matplotlib.pyplot as plt
import openpyxl

from SimPEG.config import SimpegConfig
from utils.simulation_dc_resistivity import SimulationDCResistivity

HAS_CUDA = torch.cuda.is_available()
VRAM_LIMIT_MB = 7000


# ---------------------------------------------------------------------------
# Modelo sintetico
# ---------------------------------------------------------------------------
def make_resistivity_model(nx, nz, seed=42):
    """Heterogeneous 2D resistivity model using smoothed random field.

    Generates a log-normal random field smoothed with a Gaussian kernel
    to produce spatially correlated resistivity variations (10-1000 ohm-m).
    """
    from scipy.ndimage import gaussian_filter
    rng = np.random.RandomState(seed)
    # Log-normal random field
    log_rho = rng.randn(nz, nx)
    # Smooth with Gaussian kernel (correlation length ~ 15% of model size)
    sigma = max(min(nx, nz) * 0.15, 2.0)
    log_rho = gaussian_filter(log_rho, sigma=sigma)
    # Scale to resistivity range: 10 - 1000 ohm-m
    log_rho = 1.0 + 1.5 * (log_rho - log_rho.min()) / (log_rho.max() - log_rho.min())
    rho = 10.0 ** log_rho  # range ~10 to ~316 ohm-m
    return rho


# ---------------------------------------------------------------------------
# Setup comun de simulacion
# ---------------------------------------------------------------------------
def setup_simulation(model_np, nx, nz):
    sep = 2.0
    nElec = min(nx // 2, 24)
    if nElec < 4:
        nElec = 4
    pf = nElec * sep
    nlines = min(6, nElec // 2)
    surveyinfo = dict(
        nElec=nElec, sep=sep, pi=0, depth=float(nz),
        typ_survey="dipole-dipole", dim="2D",
        data_type="volt", nlines=nlines, pf=pf
    )
    return surveyinfo


# ---------------------------------------------------------------------------
# Run: SuperLU / Pardiso / PCG-GPU
# ---------------------------------------------------------------------------
def run_dc(model_np, nx, nz, solver_name, device, nky=11):
    cfg = SimpegConfig()
    cfg.dtype = torch.float64

    if solver_name == "pcg_gpu":
        cfg.torch_is_active = True
        cfg.device = "cuda"
        cfg.solver = "superlu"
    else:
        cfg.torch_is_active = True
        cfg.device = "cpu"
        cfg.solver = solver_name

    surveyinfo = setup_simulation(model_np, nx, nz)
    m = torch.tensor(model_np.copy(), requires_grad=True, device='cpu', dtype=torch.float64)

    results = {}
    SDCR = SimulationDCResistivity(model=m, surveyinfo=surveyinfo, tensor=True)
    SDCR.initialize_survey()
    SDCR.initialize_mesh(adjust_model=False, m_background=torch.mean(m))
    results["nD"] = SDCR.survey.nD
    results["mesh_nC"] = SDCR.mesh.nC

    gc.collect()
    if device != "cpu":
        torch.cuda.empty_cache()
        torch.cuda.reset_peak_memory_stats()
    mem_before = psutil.Process().memory_info().rss

    # Forward
    t0 = time.perf_counter()
    dobs = SDCR.forward_modeling(nky=nky, error=0., only_data=True)
    if device != "cpu" and isinstance(dobs, torch.Tensor):
        torch.cuda.synchronize()
    results["fwd"] = time.perf_counter() - t0

    if device != "cpu":
        vram_used = torch.cuda.memory_allocated() / 1e6
        if vram_used > VRAM_LIMIT_MB:
            torch.cuda.empty_cache()
            raise MemoryError(f"VRAM: {vram_used:.0f}MB > {VRAM_LIMIT_MB}MB")

    # Backward
    t0 = time.perf_counter()
    if isinstance(dobs, torch.Tensor):
        dobs.sum().backward()
        if device != "cpu":
            torch.cuda.synchronize()
        results["bwd"] = time.perf_counter() - t0
    else:
        results["bwd"] = None

    results["total"] = results["fwd"] + (results["bwd"] or 0)

    if device != "cpu":
        results["mem_MB"] = torch.cuda.max_memory_allocated() / 1e6
    else:
        results["mem_MB"] = max((psutil.Process().memory_info().rss - mem_before) / 1e6, 0)

    del SDCR, dobs, m
    gc.collect()
    if device != "cpu":
        torch.cuda.empty_cache()
    return results


# ---------------------------------------------------------------------------
# Run: Dense GPU (torch.linalg.solve on dense matrix)
# ---------------------------------------------------------------------------
def run_dc_dense_gpu(model_np, nx, nz, nky=11):
    """Forward+backward usando torch.linalg.solve (densifica la matriz A)."""
    cfg = SimpegConfig()
    cfg.torch_is_active = True
    cfg.device = "cuda"
    cfg.dtype = torch.float64
    cfg.solver = "superlu"  # para setup, luego override el solve

    surveyinfo = setup_simulation(model_np, nx, nz)
    m = torch.tensor(model_np.copy(), requires_grad=True, device='cpu', dtype=torch.float64)

    SDCR = SimulationDCResistivity(model=m, surveyinfo=surveyinfo, tensor=True)
    SDCR.initialize_survey()
    SDCR.initialize_mesh(adjust_model=False, m_background=torch.mean(m))

    results = {"nD": SDCR.survey.nD, "mesh_nC": SDCR.mesh.nC}

    # Get simulation object to manually solve with dense
    from SimPEG.electromagnetics.static import resistivity as dc
    from SimPEG import maps

    sim = dc.Simulation2DNodal(
        SDCR.mesh, rhoMap=SDCR.mapping, survey=SDCR.survey,
        storeJ=False, nky=nky, torch_is_active=True
    )

    gc.collect()
    torch.cuda.empty_cache()
    torch.cuda.reset_peak_memory_stats()

    m_log = SDCR.m

    # Forward: solve each ky with dense solver
    t0 = time.perf_counter()
    sim.model = m_log
    f = sim.fieldsPair(sim)
    kys = sim._quad_points
    f._quad_weights = sim._quad_weights

    for iky, ky in enumerate(kys):
        A_sparse = sim.getA(ky)
        RHS_np = sim.getRHS(ky)

        # Densify and solve on GPU
        A_dense = A_sparse.to_dense().to("cuda")
        RHS_t = torch.tensor(RHS_np, dtype=torch.float64, device="cuda")

        # VRAM check
        vram_used = torch.cuda.memory_allocated() / 1e6
        if vram_used > VRAM_LIMIT_MB:
            del A_dense, RHS_t
            torch.cuda.empty_cache()
            raise MemoryError(f"VRAM: {vram_used:.0f}MB > {VRAM_LIMIT_MB}MB")

        u = torch.linalg.solve(A_dense, RHS_t)
        f[:, sim._solutionType, iky] = u  # mantener en CUDA (fields storage está en CUDA)
        sim.Ainv[iky] = None

        del A_dense  # liberar VRAM inmediatamente

    torch.cuda.synchronize()
    results["fwd"] = time.perf_counter() - t0

    # Backward: project fields to data, then backprop
    dobs = sim.dpred(f=f)
    t0 = time.perf_counter()
    if isinstance(dobs, torch.Tensor) and dobs.requires_grad:
        dobs.sum().backward()
        torch.cuda.synchronize()
        results["bwd"] = time.perf_counter() - t0
    else:
        results["bwd"] = None
    results["total"] = results["fwd"] + (results["bwd"] or 0)
    results["mem_MB"] = torch.cuda.max_memory_allocated() / 1e6

    del sim, SDCR, f, m
    gc.collect()
    torch.cuda.empty_cache()
    return results


# ---------------------------------------------------------------------------
# Plotting
# ---------------------------------------------------------------------------
def plot_results(all_results, solver_names, save_dir):
    params = [r["params"] for r in all_results]

    colors = {"SuperLU": "C0", "Pardiso": "C1", "PCG-GPU": "C2", "Dense-GPU": "C3"}
    markers = {"SuperLU": "o", "Pardiso": "s", "PCG-GPU": "^", "Dense-GPU": "D"}

    # --- Forward ---
    fig, ax = plt.subplots(figsize=(9, 5))
    for name in solver_names:
        x, y = [], []
        for r in all_results:
            if r.get(name) and r[name]["fwd"]:
                x.append(r["params"])
                y.append(r[name]["fwd"] * 1000)
        if x:
            ax.plot(x, y, f'-{markers.get(name,"o")}', color=colors.get(name,"gray"),
                    label=name, linewidth=2, markersize=7)
    ax.set_xlabel("Model parameters", fontsize=12)
    ax.set_ylabel("Forward time (ms)", fontsize=12)
    ax.set_title("DC Resistivity 2D - Forward Time", fontsize=14)
    ax.legend(fontsize=11)
    ax.grid(True, alpha=0.3)
    fig.tight_layout()
    fig.savefig(os.path.join(save_dir, "benchmark_forward.png"), dpi=150)
    print(f"  Saved: {os.path.join(save_dir, 'benchmark_forward.png')}")

    # --- Backward ---
    fig, ax = plt.subplots(figsize=(9, 5))
    for name in solver_names:
        x, y = [], []
        for r in all_results:
            if r.get(name) and r[name].get("bwd") and r[name]["bwd"]:
                x.append(r["params"])
                y.append(r[name]["bwd"] * 1000)
        if x:
            ax.plot(x, y, f'-{markers.get(name,"o")}', color=colors.get(name,"gray"),
                    label=name, linewidth=2, markersize=7)
    ax.set_xlabel("Model parameters", fontsize=12)
    ax.set_ylabel("Backward time (ms)", fontsize=12)
    ax.set_title("DC Resistivity 2D - Backward Time (autograd)", fontsize=14)
    ax.legend(fontsize=11)
    ax.grid(True, alpha=0.3)
    fig.tight_layout()
    fig.savefig(os.path.join(save_dir, "benchmark_backward.png"), dpi=150)
    print(f"  Saved: {os.path.join(save_dir, 'benchmark_backward.png')}")

    # --- Memory ---
    fig, ax = plt.subplots(figsize=(9, 5))
    for name in solver_names:
        x, y = [], []
        for r in all_results:
            if r.get(name) and r[name]["mem_MB"]:
                x.append(r["params"])
                y.append(r[name]["mem_MB"])
        if x:
            ax.plot(x, y, f'-{markers.get(name,"o")}', color=colors.get(name,"gray"),
                    label=name, linewidth=2, markersize=7)
    ax.set_xlabel("Model parameters", fontsize=12)
    ax.set_ylabel("Memory (MB)", fontsize=12)
    ax.set_title("DC Resistivity 2D - Memory Usage", fontsize=14)
    ax.legend(fontsize=11)
    ax.grid(True, alpha=0.3)
    fig.tight_layout()
    fig.savefig(os.path.join(save_dir, "benchmark_memory.png"), dpi=150)
    print(f"  Saved: {os.path.join(save_dir, 'benchmark_memory.png')}")

    plt.close("all")


# ---------------------------------------------------------------------------
# Export to Excel
# ---------------------------------------------------------------------------
def export_results_xlsx(all_results, solver_names, save_dir):
    wb = openpyxl.Workbook()

    # --- Sheet 1: Forward time (ms) ---
    ws_fwd = wb.active
    ws_fwd.title = "Forward (ms)"
    ws_fwd.append(["Params"] + solver_names)
    for r in all_results:
        row = [r["params"]]
        for name in solver_names:
            v = r.get(name)
            row.append(round(v["fwd"] * 1000, 2) if v and v.get("fwd") else None)
        ws_fwd.append(row)

    # --- Sheet 2: Backward time (ms) ---
    ws_bwd = wb.create_sheet("Backward (ms)")
    ws_bwd.append(["Params"] + solver_names)
    for r in all_results:
        row = [r["params"]]
        for name in solver_names:
            v = r.get(name)
            row.append(round(v["bwd"] * 1000, 2) if v and v.get("bwd") else None)
        ws_bwd.append(row)

    # --- Sheet 3: Total time (ms) ---
    ws_tot = wb.create_sheet("Total (ms)")
    ws_tot.append(["Params"] + solver_names)
    for r in all_results:
        row = [r["params"]]
        for name in solver_names:
            v = r.get(name)
            row.append(round(v["total"] * 1000, 2) if v and v.get("total") else None)
        ws_tot.append(row)

    # --- Sheet 4: Memory (MB) ---
    ws_mem = wb.create_sheet("Memory (MB)")
    ws_mem.append(["Params"] + solver_names)
    for r in all_results:
        row = [r["params"]]
        for name in solver_names:
            v = r.get(name)
            row.append(round(v["mem_MB"], 2) if v and v.get("mem_MB") is not None else None)
        ws_mem.append(row)

    # --- Sheet 5: Speedup vs SuperLU ---
    ws_sp = wb.create_sheet("Speedup vs SuperLU")
    ws_sp.append(["Params"] + solver_names)
    for r in all_results:
        base = r.get("SuperLU")
        base_t = base["total"] if base else None
        row = [r["params"]]
        for name in solver_names:
            v = r.get(name)
            if v and base_t and v["total"] > 0:
                row.append(round(base_t / v["total"], 3))
            else:
                row.append(None)
        ws_sp.append(row)

    fpath = os.path.join(save_dir, "benchmark_results.xlsx")
    wb.save(fpath)
    print(f"  Saved: {fpath}")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def run():
    MODEL_CONFIGS = [
        (10, 10),       # 100
        (25, 20),       # 500
        (50, 20),       # 1000
        (67, 30),       # ~2000
        (75, 40),       # 3000
        (80, 50),       # 4000
        (100, 50),      # 5000
        (100, 60),      # 6000
        (100, 80),      # 8000
        (100, 90),      # 9000
        (100, 100),     # 10000
    ]

    solvers = [
        ("SuperLU",  "superlu",  "cpu"),
        ("Pardiso",  "pardiso",  "cpu"),
    ]
    if HAS_CUDA:
        solvers.append(("PCG-GPU", "pcg_gpu", "cuda"))

    solver_names = [s[0] for s in solvers]
    if HAS_CUDA:
        solver_names.append("Dense-GPU")

    print("=" * 110)
    print("  BENCHMARK: DC Resistivity 2D - Forward + Backward")
    print("=" * 110)
    if HAS_CUDA:
        print(f"  GPU:    {torch.cuda.get_device_name(0)}")
        props = torch.cuda.get_device_properties(0)
        vram = getattr(props, 'total_mem', None) or props.total_memory
        print(f"  VRAM:   {vram / 1e9:.1f} GB (limit: {VRAM_LIMIT_MB}MB)")
    print(f"  Torch:  {torch.__version__}")
    print(f"  nky:    11")
    print()

    # Header
    print(f"  {'params':>7}", end="")
    for name in solver_names:
        print(f" | {name+' fwd':>10} {name+' bwd':>10} {'mem':>8}", end="")
    print()
    print("  " + "-" * (9 + 32 * len(solver_names)))

    all_results = []

    for nx, nz in MODEL_CONFIGS:
        n_params = nx * nz
        model_np = make_resistivity_model(nx, nz)

        row = {"params": n_params}
        line = f"  {n_params:>7}"

        # Standard solvers
        for label, solver_name, device in solvers:
            try:
                r = run_dc(model_np, nx, nz, solver_name, device)
                bwd_str = f"{r['bwd']*1000:>9.0f}ms" if r.get('bwd') else "      N/A"
                line += f" | {r['fwd']*1000:>9.0f}ms {bwd_str} {r['mem_MB']:>7.0f}MB"
                row[label] = r
            except Exception as e:
                line += f" | {'ERR':>10} {'ERR':>10} {'ERR':>8}"
                row[label] = None
                print(f"\n    [{label} {n_params}] {e}\n")

        # Dense GPU
        if HAS_CUDA:
            try:
                r = run_dc_dense_gpu(model_np, nx, nz)
                bwd_str = f"{r['bwd']*1000:>9.0f}ms" if r.get('bwd') else "      N/A"
                line += f" | {r['fwd']*1000:>9.0f}ms {bwd_str} {r['mem_MB']:>7.0f}MB"
                row["Dense-GPU"] = r
            except Exception as e:
                line += f" | {'ERR':>10} {'---':>10} {'ERR':>8}"
                row["Dense-GPU"] = None
                print(f"\n    [Dense-GPU {n_params}] {e}\n")

        print(line)
        all_results.append(row)
        gc.collect()

    # Speedup summary
    print()
    print("=" * 110)
    print("  SPEEDUP vs SuperLU (total)")
    print("=" * 110)
    print(f"  {'params':>7}", end="")
    for name in solver_names:
        print(f"  {name:>14}", end="")
    print()

    for row in all_results:
        base = row.get("SuperLU")
        base_t = base["total"] if base else None
        line = f"  {row['params']:>7}"
        for name in solver_names:
            r = row.get(name)
            if r and base_t and r["total"] > 0:
                line += f"  {base_t / r['total']:>12.2f}x "
            else:
                line += f"  {'N/A':>14}"
        print(line)

    # Plots
    print()
    save_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)))
    plot_results(all_results, solver_names, save_dir)

    # Export to Excel
    export_results_xlsx(all_results, solver_names, save_dir)
    print("\nDone.")


if __name__ == "__main__":
    run()
